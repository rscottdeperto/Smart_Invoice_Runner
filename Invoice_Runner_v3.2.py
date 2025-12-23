# -*- coding: utf-8 -*-
"""
Smart Invoice Runner – v3.2 (2025-12-10)

WHAT THIS APP DOES
------------------
• Auto-detects FedEx invoices and parses them locally (PyMuPDF) — no API calls.
• Auto-detects Lightning Messenger Express invoices and parses them locally (PyMuPDF) — no API calls.
• For all other vendors, All invoices are processed locally. No cloud or Azure calls are made.
• FedEx output:
  - Sets Description="FedEx" for shipment rows and "FedEx Other Charges" for that row
  - Adds explicit columns: FedEx_Sender, FedEx_CustRef, PrimaryClientCode
  - (Tracking/Order# removed from the output as requested)
• Lightning Messenger output (per Reference):
  - InvoiceDate := first Order "Date" under that Reference (MM/DD/YYYY)
  - (Tracking/OrderID previously captured as Order ID is no longer emitted)
  - FedEx_Sender (labeled "Caller/Sender") := the Caller captured near that Order ID
  - FedEx_CustRef (labeled "Reference") := the "Billing Reference 1" number
  - Amount := the per-Reference Total (from "Totals: Billing Reference..." lines)
• UI improvements:
  - Splash popup on launch that shows Sophos Connect image
  - Status bubble “Analyzing… Please wait” while processing after clicking Analyze
  - Confirmation popup “Load Complete” after the Client Code Map (CSV) is loaded
  - Graphical progress bar while processing files
  - Client Code Map (CSV): browse & load to map 10-digit CustRef to PrimaryClientCode
  - Mixed-folder processing: drop in a folder with FedEx + Lightning + other PDFs and click Analyze

UNIFIED OUTPUT COLUMNS (both pipelines)
---------------------------------------
InvoiceFileName
Vendor
InvoiceID
InvoiceDate
DueDate
Description
Quantity
UnitPrice
Amount
Currency
FedEx_Sender (UI label: "Caller/Sender")
FedEx_CustRef (UI label: "Reference")
PrimaryClientCode

© Gelfand, Rennert & Feldman, LLC
"""
from io import BytesIO
import fitz  # PyMuPDF
import re
import csv
import time
import requests  # pip install request
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Tuple

# ---------- UI ----------
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
try:
    import customtkinter as ctk  # pip install customtkinter
    USE_CTK = True
except Exception:
    USE_CTK = False

from datetime import datetime


def normalize_date(date_str):
    """
    Attempts to parse a date string in various formats and returns it as YYYY-MM-DD.
    If parsing fails, returns the original string.
    """
    date_str = date_str.strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except Exception:
            continue
    # Try to parse date ranges like "10/16/2025-10/31/2025" (take the last date)
    if "-" in date_str:
        parts = [p.strip() for p in date_str.split("-")]
        for part in reversed(parts):
            norm = normalize_date(part)
            if norm != part:
                return norm
    return date_str


# ---------- Local PDF text extraction ----------
# pip install pymupdf

# ---------- Splash image helpers ----------
try:
    from PIL import Image, ImageTk  # pip install pillow
    PIL_OK = True
except Exception:
    PIL_OK = False

# --- OCR imports (optional; only used if a page is image-only) ---
try:
    from pdf2image import convert_from_path
    import pytesseract
    OCR_AVAILABLE = True
except Exception:
    OCR_AVAILABLE = False


APP_VERSION = "SmartInvoiceRunner v3.2"
MAX_FILE_MB = 50
SPLASH_IMAGE_URL = r"C:\Users\rscottdeperto\Desktop\Invoice Testing\Coding\assets\splash.png"

# ======================================
# Utilities
# ======================================


def normalize_text(text: str) -> str:
    if not text:
        return ""
    text = text.replace("\xa0", " ").replace(
        "\u200b", " ").replace("\ufb01", "fi")
    text = re.sub(r"[\t ]+", " ", text)
    # Normalize both old and new FedEx total charge labels to a single string
    text = re.sub(r"Total\s*Transportation\s*Charges",
                  "Total Charge", text, flags=re.I)
    text = re.sub(r"Total\s*Charge", "Total Charge", text, flags=re.I)
    # Lightning: normalize hyphens/dashes around "Totals: Billing Reference"
    text = re.sub(
        r"Totals:\s*Billing\s*Reference\s*1\s*[\u2013\-]\s*",
        "Totals: Billing Reference 1 - ",
        text,
        flags=re.I
    )
    # Lightning: aggressively normalize OCR variations for all key anchors
    text = re.sub(r"summary\s*[-–—]?\s*billing\s*reference\s*1",
                  "Summary - Billing Reference 1", text, flags=re.I)
    text = re.sub(r"totals?:\s*billing\s*reference\s*1",
                  "Totals: Billing Reference 1", text, flags=re.I)
    text = re.sub(r"order\s*total\s*:", "Order Total:", text, flags=re.I)
    text = re.sub(r"billing\s*reference\s*1",
                  "Billing Reference 1", text, flags=re.I)
    text = re.sub(r"customer\s*number", "Customer Number", text, flags=re.I)
    text = re.sub(r"invoice\s*number", "Invoice Number", text, flags=re.I)
    text = re.sub(r"invoice\s*period", "Invoice Period", text, flags=re.I)
    return text


def _ocr_pdf_to_text(pdf_path: Path, dpi: int = 300) -> str:
    """
    Fallback OCR using pdf2image + Tesseract. Returns concatenated text for all pages.
    """
    if not OCR_AVAILABLE:
        return ""
    try:
        images = convert_from_path(str(pdf_path), dpi=dpi)
        out_parts = []
        for img in images:
            # Tesseract English; adjust if needed
            out_parts.append(pytesseract.image_to_string(img, lang='eng'))
        return "\n".join(out_parts)
    except Exception:
        return ""


def _has_meaningful_text(text: str, min_len: int = 60) -> bool:
    """
    Heuristic to decide if the extracted text is sufficient
    (avoids OCR for already-text PDFs).
    """
    return len((text or "").strip()) >= min_len


def read_pdf_text(file_path: Path) -> str:
    """
    Extracts text with PyMuPDF first. If little/no text is found (likely scanned),
    falls back to OCR (pdf2image + Tesseract). Output is normalized for parsers.
    """
    # First pass: native text extraction
    parts = []
    try:
        with fitz.open(str(file_path)) as doc:
            for pg in doc:
                # 'text' for layout-friendly content; switch to 'plain' if needed
                parts.append(pg.get_text("text"))
    except Exception:
        parts = []

    raw = "\n".join(parts)
    if _has_meaningful_text(raw):
        return normalize_text(raw)

    # Fallback: OCR (only if libraries are present)
    ocr_text = _ocr_pdf_to_text(file_path, dpi=300)
    if _has_meaningful_text(ocr_text):
        return normalize_text(ocr_text)

    # If OCR not available or failed, return normalized (possibly empty) native text
    return normalize_text(raw)


def soft_clean(s: Optional[str]) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def try_parse_date(human_date: str) -> str:
    for fmt in ("%b %d, %Y", "%B %d, %Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(human_date.strip(), fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return human_date.strip()


def amount_to_float(s: str) -> Optional[float]:
    if s is None:
        return None
    try:
        return float(str(s).replace(",", "").replace("$", ""))
    except Exception:
        return None


def map_primary_from_custref(cust: str, client_map: Optional[Dict[str, str]]) -> str:
    """
    Looks up the PrimaryClientCode for any reference string (any length, any format).
    - Tries exact match (case-sensitive and insensitive)
    - Tries substring match (case-insensitive)
    Returns "" if no match or map.
    """
    if not cust or not client_map:
        return ""
    ref = soft_clean(str(cust))
    # 1. Exact match
    if ref in client_map:
        return client_map[ref]
    # 2. Case-insensitive match
    for k, v in client_map.items():
        if ref.lower() == k.lower():
            return v
    # 3. Substring match (case-insensitive)
    for k, v in client_map.items():
        if k and k.lower() in ref.lower():
            return v
    return ""

# ======================================
# FedEx Local Parser (page-break aware)
# ======================================


class FedExParser:
    INVOICE_HEADER_RX = re.compile(
        r"Invoice\s+Number\s+([^\s]+).*?Invoice\s+Date\s+([A-Za-z]{3,9}\s+\d{1,2},\s+\d{4})",
        re.I | re.S
    )

    TOTAL_RX = re.compile(
        r"Total\s*(?:Charge|Transportation\s*Charges)\s+USD\s+\$?\s*([\d,]+\.?\d{2})",
        re.I
    )

    # used for merge only; not emitted
    TRACK_RX = re.compile(r"Tracking\s*ID[:\s]+(\d{9,18})", re.I)
    CONT_RX = re.compile(
        r"(continued\s+on\s+next\s+page\nTracking\s*ID[:\s]+\d+\s+continued)", re.I
    )
    CUST_REF_RX = re.compile(r"Cust\.\s*Ref\.?\s*:\s*(.+)", re.I)
    SENDER_RX = re.compile(r"^\s*Sender\s+(.+)$", re.I | re.M)

    OTHER_CHARGES_RX = re.compile(
        r"Other Charges\s*USD\s*\$?\s*([\d,.]+)", re.I
    )
    # Also match "Late Fee" as Other Charges
    LATE_FEE_RX = re.compile(
        r"Late Fee.*?(\d{2}/\d{2}/\d{2,4}).*?([\d,.]+)$", re.I | re.M)

    USD_RX = re.compile(r"\bUSD\b", re.I)

    def __init__(self, client_map: Optional[Dict[str, str]] = None):
        self.client_map = client_map or {}

    @staticmethod
    def iter_ship_blocks(text: str, start_label="Ship Date:"):
        starts = [m.start() for m in re.finditer(re.escape(start_label), text)]
        if not starts:
            return
        starts.append(len(text))
        for i in range(len(starts) - 1):
            yield text[starts[i]:starts[i + 1]]

    @staticmethod
    def extract_sender_name(sender_line: str) -> str:
        s = re.sub(r"^\s*Sender\s+", "", sender_line, flags=re.I).strip()
        cut = re.split(r"\bGelfand\b", s, flags=re.I)[0].strip()
        if cut:
            cut = re.sub(r"\s+Gelfand.*$", "", cut, flags=re.I).strip()
            return cut[:80]
        tokens = s.split()
        return (" ".join(tokens[:3]))[:80] if tokens else s[:80]

    def parse_invoice_header(self, text: str) -> Tuple[Optional[str], Optional[str]]:
        m = self.INVOICE_HEADER_RX.search(
            text[:4000] if len(text) > 4000 else text)
        if not m:
            return None, None
        inv_no = m.group(1).strip()
        inv_date_iso = try_parse_date(m.group(2).strip())
        return inv_no, inv_date_iso

    def parse_other_charges(self, text: str) -> Optional[float]:
        # Try "Other Charges" summary
        m = self.OTHER_CHARGES_RX.search(text)
        if m:
            return amount_to_float(m.group(1))
        # Try "Late Fee" detail
        m2 = self.LATE_FEE_RX.search(text)
        if m2:
            return amount_to_float(m2.group(2))

    def is_usd(self, text: str) -> bool:
        return bool(self.USD_RX.search(text))

    def parse(self, pdf_text: str, file_name: str) -> List[Dict]:
        """Returns a list of unified output rows for FedEx."""
        inv_no, inv_date = self.parse_invoice_header(pdf_text)
        currency = "USD" if self.is_usd(pdf_text) else ""
        rows: List[Dict] = []
        pending = None  # {"tracking","cust","sender"}

        def emit_row(sender, cust, tracking, total_amt):
            primary = map_primary_from_custref(
                soft_clean(cust or ""), self.client_map)
            rows.append({
                "InvoiceFileName": file_name,
                "Vendor": "FedEx",
                "InvoiceID": inv_no or "",
                "InvoiceDate": inv_date or "",
                "DueDate": "",
                "Description": "FedEx",
                "Quantity": "",
                "UnitPrice": "",
                "Amount": total_amt if total_amt is not None else "",
                "Currency": currency,
                "FedEx_Sender": sender or "",
                "FedEx_CustRef": soft_clean(cust or ""),
                "PrimaryClientCode": primary or "",
            })

        # Walk all "Ship Date:" blocks and merge page-break splits
        for blk in self.iter_ship_blocks(pdf_text, "Ship Date:"):
            # Fields
            mref = self.CUST_REF_RX.search(blk)
            cust_line = mref.group(1).strip() if mref else None
            msend = self.SENDER_RX.search(blk)
            sender = self.extract_sender_name(
                msend.group(0)) if msend else None
            totals = list(self.TOTAL_RX.finditer(blk))
            total_amt = amount_to_float(
                totals[-1].group(1)) if totals else None
            mtrk = self.TRACK_RX.search(blk)
            tracking = mtrk.group(1).strip() if mtrk else None
            continued = bool(self.CONT_RX.search(blk))

            if pending:
                same = tracking and pending.get(
                    "tracking") and tracking == pending["tracking"]
                if total_amt is not None and (same or continued or not tracking):
                    emit_row(pending.get("sender") or sender,
                             pending.get("cust") or cust_line,
                             pending.get("tracking") or tracking,
                             total_amt)
                    pending = None
                    continue
                if same:
                    if not pending.get("cust"):
                        pending["cust"] = cust_line
                    if not pending.get("sender"):
                        pending["sender"] = sender
                    continue
                pending = None  # different shipment began; drop incomplete pending

            if total_amt is not None:
                emit_row(sender, cust_line, tracking, total_amt)
            else:
                if sender or cust_line or continued or tracking:
                    pending = {"tracking": tracking,
                               "cust": cust_line, "sender": sender}

        # Add Other Charges as its own line
        oc = self.parse_other_charges(pdf_text)
        if oc is not None:
            rows.append({
                "InvoiceFileName": file_name,
                "Vendor": "FedEx",
                "InvoiceID": inv_no or "",
                "InvoiceDate": inv_date or "",
                "DueDate": "",
                "Description": "FedEx Other Charges",
                "Quantity": "",
                "UnitPrice": "",
                "Amount": oc,
                "Currency": "USD" if currency == "USD" or "USD" in pdf_text else "",
                "FedEx_Sender": "",
                "FedEx_CustRef": "",
                "PrimaryClientCode": "",
            })

        # No deduplication: allow all rows, including duplicates
        return rows

# ======================================
# Lightning Messenger (local) Parser
# ======================================


class LightningParser:
    """
    Parses Lightning Messenger Express invoices from text.
    Produces ONE row per Reference (Billing Reference 1).
    """
    # Header variants (capture Invoice Number and Invoice Date)
    HDR_A = re.compile(
        r"Invoice\s+Number\s+(\d{3,})\s+(\d{1,2}/\d{1,2}/\d{4})", re.I)
    HDR_B = re.compile(
        r"Customer\s+Number\s+Invoice\s+Number\s+Invoice\s+Date\s+Invoice\s+Amount\s+\d+\s+(\d{3,})\s+(\d{1,2}/\d{1,2}/\d{4})",
        re.I
    )
    # Reference totals across pages:
    # "Totals: Billing Reference 1 - 3119952000 Total: $35.00"
    REF_TOTAL_RX = re.compile(
        r"Totals:\s*Billing\s*Reference\s*1\s*\-\s*(\d{7,})\s*Total:\s*\$?\s*([\d,]+\.\d{2})",
        re.I
    )
    # Find each "Billing Reference 1 <ref>" section boundary
    REF_START_RX = re.compile(r"Billing\s+Reference\s+1\s+(\d{7,})", re.I)
    # Within a reference block, find the first date (MM/DD/YYYY)
    DATE_RX = re.compile(r"\b(\d{1,2}/\d{1,2}/\d{4})\b")
    # After "Order ID" label, next numeric token is the Order ID (e.g., 589396.01)
    ORDER_ID_FROM_LABEL = re.compile(
        r"Order\s+ID\s+([0-9]{5,}(?:\.\d{2})?)", re.I)
    # Fallback: numeric token following the date, limited window
    ORDER_ID_FALLBACK = re.compile(r"\b([0-9]{5,}(?:\.\d{2})?)\b")

    def __init__(self, client_map: Optional[Dict[str, str]] = None):
        self.client_map = client_map or {}

    @staticmethod
    def _extract_header(text: str) -> Tuple[str, str]:
        for rx in (LightningParser.HDR_A, LightningParser.HDR_B):
            m = rx.search(text[:5000])
            if m:
                inv_no, inv_date = m.group(1), m.group(2)
                return inv_no, try_parse_date(inv_date)
        return "", ""

    @staticmethod
    def _totals_by_reference(text: str) -> Dict[str, float]:
        out: Dict[str, float] = {}
        for ref, amt in LightningParser.REF_TOTAL_RX.findall(text):
            out[ref] = amount_to_float(amt) or 0.0
        return out

    @staticmethod
    def _caller_near(order_id: str, block: str) -> str:
        """
        Heuristic: capture tokens immediately after the order_id up to 'Gelfand' or a newline,
        which tends to be the Caller. Keeps phones if present (e.g., "Marine 310-282-5973").
        """
        if not order_id:
            return ""
        m = re.search(re.escape(order_id) + r"\s+(.{2,100}?)\s+(?:Gelfand|SB|City\s+of|Deliver|-|\d{3,})",
                      block, re.S | re.I)
        if m:
            return soft_clean(m.group(1))
        # Fallback: take next 40 chars
        m2 = re.search(re.escape(order_id) +
                       r"\s+(.{2,60})", block, re.S | re.I)
        return soft_clean(m2.group(1)) if m2 else ""

    def extract_lightning_orders(self, text, file_name):
        """
        Robustly extracts Lightning Messenger order lines from OCR'd or digital text.
        Returns a list of dicts, one per order.
        """
        orders = []
        # Normalize text: collapse multiple spaces, fix common OCR issues
        lines = [re.sub(r'\s+', ' ', l).strip()
                 for l in text.split('\n') if l.strip()]
        block = {}
        for line in lines:
            # Billing Reference
            m_ref = re.search(
                r'Billing Reference 1\s*[-:]?\s*([\w\- ]+)', line, re.I)
            if m_ref:
                if block:  # Save previous block if exists
                    orders.append(block)
                    block = {}
                block['Billing Reference'] = m_ref.group(1).strip()
            # Order ID
            m_oid = re.search(r'Order ID\s*([0-9.]+)', line, re.I)
            if m_oid:
                block['Order ID'] = m_oid.group(1).strip()
            # Caller
            m_caller = re.search(r'Caller\s*([A-Za-z .]+)', line, re.I)
            if m_caller:
                block['Caller'] = m_caller.group(1).strip()
            # Order Total
            m_total = re.search(
                r'Order Total[: ]*\$?([\d,]+\.\d{2})', line, re.I)
            if m_total:
                block['Order Total'] = m_total.group(1).replace(',', '')
            # Origin
            m_origin = re.search(r'Origin\s*([A-Za-z0-9 ,&\-.]+)', line, re.I)
            if m_origin:
                block['Origin'] = m_origin.group(1).strip()
            # Destination
            m_dest = re.search(
                r'Destination\s*([A-Za-z0-9 ,&\-.]+)', line, re.I)
            if m_dest:
                block['Destination'] = m_dest.group(1).strip()
            # If we reach a "Totals:" line, treat as end of block
            if re.search(r'Totals?:', line, re.I) and block:
                orders.append(block)
                block = {}
        # Catch any trailing block
        if block:
            orders.append(block)
        # Add file name to each order
        for o in orders:
            o['InvoiceFileName'] = file_name
        return orders

    def parse(self, pdf_text: str, file_name: str) -> List[Dict]:
        rows: List[Dict] = []
        inv_no, inv_date_hdr = self._extract_header(pdf_text)
        totals_map = self._totals_by_reference(pdf_text)

        # Find all reference start positions
        refs = [(m.group(1), m.start())
                for m in self.REF_START_RX.finditer(pdf_text)]
        if not refs:
            return rows

        # Add sentinel end
        refs_with_end = []
        for i, (ref, start) in enumerate(refs):
            end = refs[i + 1][1] if i + 1 < len(refs) else len(pdf_text)
            refs_with_end.append((ref, start, end))

        for ref, start, end in refs_with_end:
            block = pdf_text[start:end]
            # Date: first date occurrence inside the block
            date_found = ""
            m_date = self.DATE_RX.search(block)
            if m_date:
                date_found = try_parse_date(m_date.group(1))

            # Order ID (no longer emitted, used to find Caller)
            order_id = ""
            m_oid = self.ORDER_ID_FROM_LABEL.search(block)
            if m_oid:
                order_id = m_oid.group(1)
            else:
                if m_date:
                    m_fallback = self.ORDER_ID_FALLBACK.search(
                        block[m_date.end(): m_date.end() + 160]
                    )
                    if m_fallback:
                        order_id = m_fallback.group(1)

            # Caller near order id
            caller = self._caller_near(order_id, block) if order_id else ""

            # Total for this Reference
            amt = totals_map.get(ref, None)

            # Map PrimaryClientCode from CustRef (ref)
            primary = map_primary_from_custref(
                soft_clean(ref or ""), self.client_map)

            rows.append({
                "InvoiceFileName": file_name,
                "Vendor": "Lightning Messenger Express",
                "InvoiceID": inv_no,
                "InvoiceDate": date_found or inv_date_hdr,  # normalize Date into InvoiceDate
                "DueDate": "",
                "Description": "Lightning Messenger",
                "Quantity": "",
                "UnitPrice": "",
                "Amount": amt if amt is not None else "",
                "Currency": "USD",
                "FedEx_Sender": caller,   # label: Caller/Sender
                "FedEx_CustRef": ref,     # label: Reference
                "PrimaryClientCode": primary or "",
            })

        return rows

# ======================================
# Generic "Other" Parser
# ======================================


def generic_invoice_parser(pdf_text: str, file_name: str) -> list:
    """
    Attempts to extract basic invoice fields from any vendor's invoice.
    Returns a list with one dict per invoice found (usually one per file).
    """
    # Patterns for common fields
    patterns = {
        "InvoiceID": [
            r"Invoice\s*Number[:\s]*([A-Z0-9\-]+)",
            r"Inv(?:oice)?\s*#[:\s]*([A-Z0-9\-]+)"
        ],
        "InvoiceDate": [
            r"Invoice\s*Date[:\s]*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})",
            r"Date of issue[:\s]*([A-Za-z]+\s+\d{1,2},\s*\d{4})",
            r"Date due[:\s]*([A-Za-z]+\s+\d{1,2},\s*\d{4})",
            r"Invoice Period[:\s]*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4}(?:\s*-\s*[0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})?)"
        ],
        "Amount": [
            r"Total\s*Amount[:\s\$]*([0-9,]+\.\d{2})",
            r"Amount\s*Due[:\s\$]*([0-9,]+\.\d{2})",
            r"Total[:\s\$]*([0-9,]+\.\d{2})",
            r"Amount due\s*\$?([0-9,]+\.\d{2})"
        ]
    }

    # Try to extract each field
    result = {
        "InvoiceFileName": file_name,
        "Vendor": "",
        "InvoiceID": "",
        "InvoiceDate": "",
        "DueDate": "",
        "Description": "Generic Invoice",
        "Quantity": "",
        "UnitPrice": "",
        "Amount": "",
        "Currency": "",
        "FedEx_Sender": "",
        "FedEx_CustRef": "",
        "PrimaryClientCode": ""
    }

    for field, regexes in patterns.items():
        for rx in regexes:
            m = re.search(rx, pdf_text, re.IGNORECASE)
            if m:
                val = m.group(1).strip()
                if field in ("InvoiceDate", "DueDate"):
                    val = normalize_date(val)
                result[field] = val
                break

    # ...and for DueDate extraction...
    due_date_match = re.search(
        r"Date due[:\s]*([A-Za-z]+\s+\d{1,2},\s*\d{4}|[0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})", pdf_text, re.IGNORECASE)
    if due_date_match:
        result["DueDate"] = normalize_date(due_date_match.group(1).strip())

    # --- Vendor Extraction ---
    # Try to find a vendor name by looking for lines near "Invoice", "From", "Bill to", or "Remit Payment To"
    lines = [line.strip() for line in pdf_text.split('\n') if line.strip()]
    vendor_candidates = []
    for i, line in enumerate(lines):
        # Look for lines that are likely to be company names
        if re.search(r'(Inc\.|LLC|Ltd|Company|Corporation|Collective|Express|Chartmetric)', line, re.IGNORECASE):
            vendor_candidates.append(line)
        # Look for "Remit Payment To" or "From" or "Bill to"
        if re.search(r'(Remit Payment To|From|Bill to|Vendor)', line, re.IGNORECASE):
            # Next non-empty line is likely the vendor
            for j in range(i+1, min(i+4, len(lines))):
                if lines[j] and not re.search(r'Invoice|Date|Number|Amount|Bill to|Ship to|Due', lines[j], re.IGNORECASE):
                    vendor_candidates.append(lines[j])
                    break
    # Remove duplicates, prefer the first
    vendor_candidates = list(dict.fromkeys(vendor_candidates))
    if vendor_candidates:
        result["Vendor"] = vendor_candidates[0]
    else:
        # Fallback: look for a company name at the top
        for line in lines[:6]:
            if re.search(r'(Inc\.|LLC|Ltd|Company|Corporation|Collective|Express|Chartmetric)', line, re.IGNORECASE):
                result["Vendor"] = line
                break

    # --- Due Date Extraction ---
    due_date_match = re.search(
        r"Date due[:\s]*([A-Za-z]+\s+\d{1,2},\s*\d{4}|[0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})", pdf_text, re.IGNORECASE)
    if due_date_match:
        result["DueDate"] = due_date_match.group(1).strip()

    # --- Currency Extraction ---
    currency_match = re.search(
        r"\b(USD|EUR|GBP|AUD|CAD|JPY|CHF|CNY|INR)\b", pdf_text)
    if currency_match:
        result["Currency"] = currency_match.group(1)

    return [result]


# ======================================
# Vendor Detection
# ======================================


def looks_like_fedex(pdf_text: str) -> bool:
    """
    OCR-tolerant FedEx detection:
    - Brand 'FedEx' + any of several shipment anchors.
    """
    if not pdf_text:
        return False
    lt = pdf_text.lower()

    if "fedex" not in lt:
        return False

    anchors = [
        "tracking id",
        "fedex express shipment", "fedex express ship",
        "ship date:", "transportation charge", "total transportation charges",
        "fedex other charges", "earned discount", "fuel surcharge",
        "invoice summary", "total this invoice"
    ]
    return any(a in lt for a in anchors)


def looks_like_lightning(pdf_text: str) -> bool:
    """
    OCR-tolerant Lightning Messenger Express detection:
    - Brand + structure anchors that survive OCR.
    """
    if not pdf_text:
        return False
    lt = pdf_text.lower()

    brand_hits = any(h in lt for h in [
        "lightning messenger express",
        "www.lightningmessengerexpress.com",
        "payment due upon receipt",
    ])

    structure_hits = sum(a in lt for a in [
        "summary - billing reference 1",
        "billing reference 1",
        "order total:",
        "totals: billing reference 1",
        "customer number",
        "invoice number",
        "invoice period",
    ])

    # Brand OR (>=2 structural anchors) is enough
    return brand_hits or structure_hits >= 2

# ======================================
# (Optional) helper if you want a single call
# ======================================


def process_file_auto(file_path: Path,
                      client_map: Optional[Dict[str, str]] = None) -> List[Dict]:
    # Skip files that are too large
    if file_path.stat().st_size / (1024 * 1024) > MAX_FILE_MB:
        return []

    txt = read_pdf_text(file_path)
    lt = (txt or "").lower()

    # Try FedEx parser first
    if looks_like_fedex(txt):
        fedex = FedExParser(client_map=client_map)
        return fedex.parse(txt, file_path.name)

    # Then try Lightning parser
    if looks_like_lightning(txt):
        lightning = LightningParser(client_map=client_map)
        return lightning.parse(txt, file_path.name)

    # If not recognized, try both parsers and return whichever yields more rows
    fedex_rows = FedExParser(client_map=client_map).parse(txt, file_path.name)
    lightning_rows = LightningParser(
        client_map=client_map).parse(txt, file_path.name)
    if len(lightning_rows) >= len(fedex_rows) and len(lightning_rows) > 0:
        return lightning_rows
    if len(fedex_rows) > 0:
        return fedex_rows

    # If still nothing, return an empty list (or optionally log for review)
    if len(lightning_rows) >= len(fedex_rows) and len(lightning_rows) > 0:
        return lightning_rows
    if len(fedex_rows) > 0:
        return fedex_rows

    # If still nothing, return a generic row
    return generic_invoice_parser(txt, file_path.name)

    # ===== SECONDARY HEURISTICS: weak but indicative anchors =====
    fedex_weak = any(a in lt for a in [
        "tracking id", "transportation charge", "earned discount", "fuel surcharge"
    ])
    lightning_weak = any(a in lt for a in [
        "billing reference 1", "order total:", "totals: billing reference 1"
    ])

    # Bias toward local if weak signals present
    if fedex_weak and not lightning_weak:
        fedex = FedExParser(client_map=client_map)
        return fedex.parse(txt, file_path.name)

    if lightning_weak and not fedex_weak:
        lightning = LightningParser()
        return lightning.parse(txt, file_path.name)


# ======================================
# Unified UI
# ======================================
COLUMNS_UNIFIED = (
    "InvoiceFileName", "Vendor", "InvoiceID", "InvoiceDate", "DueDate",
    "Description", "Quantity", "UnitPrice", "Amount", "Currency",
    "FedEx_Sender", "FedEx_CustRef", "PrimaryClientCode"
)

# Display labels for selected columns (UI headers + export header row)
COLUMN_LABELS = {
    "FedEx_Sender": "Caller/Sender",
    "FedEx_CustRef": "Reference",
    # Leave others as-is
}


def display_label(col_key: str) -> str:
    return COLUMN_LABELS.get(col_key, col_key)


def instructions_text() -> str:
    return (
        "How to use this tool\n"
        "1) Select a PDF file OR a folder of PDFs and click Analyze.\n"
        " • FedEx & Lightning Messenger Express invoices are parsed locally.\n"
        " • All other vendors use generic python scripts.\n"
        "2) (Optional) Load a Client Code Map (CSV) to populate PrimaryClientCode for FedEx.\n"
        "3) Click Analyze again. The table will populate with rows.\n"
        "4) Export to Excel or CSV using the buttons above the table.\n\n"
        "Notes\n"
        "• FedEx rows set Description=\"FedEx\" and include Caller/Sender, Reference, PrimaryClientCode.\n"
        "• Lightning rows set Description=\"Lightning Messenger\" and include Caller/Sender and Reference; "
        "Date is stored in InvoiceDate and Amount is the Reference Total.\n"
        "• Mixed folders are supported; the app routes each invoice automatically."
    )


class AppBase:
    def __init__(self):
        self.rows: List[Dict] = []
        self.columns = COLUMNS_UNIFIED
        self.client_map: Dict[str, str] = {}
        # overlays
        self._status_bubble = None
        self._splash = None

    # Abstracts for bridge
    def set_status(self, msg: str): raise NotImplementedError
    def clear_table(self): raise NotImplementedError
    def add_row(self, values: List): raise NotImplementedError
    def rebuild_tree(self): raise NotImplementedError
    def get_tree_column_width(self, name): raise NotImplementedError
    def set_tree_rowheight(self, px): raise NotImplementedError
    def after_call(self, ms, func): raise NotImplementedError
    def set_progress(self, done: int, total: int): pass  # optional override
    # overlays (to be implemented)
    def show_status_bubble(self, msg: str): pass
    def hide_status_bubble(self): pass
    def show_launch_splash(self): pass
    def hide_launch_splash(self): pass

    # ---- Client Code Map (CSV)
    def load_client_map_csv(self, csv_path: Path) -> Tuple[int, int]:
        """
        Reads CSV and builds a mapping dict. Prefers columns named
        'CustRef' and 'PrimaryClientCode'; otherwise uses the first two columns.
        Returns (rows_read, pairs_mapped).
        """
        self.client_map.clear()
        read = mapped = 0
        with csv_path.open("r", newline="", encoding="utf-8") as f:
            rdr = csv.DictReader(f)
            if not rdr.fieldnames or len(rdr.fieldnames) < 2:
                raise ValueError("CSV must have at least two columns.")
            cols = {k.lower(): k for k in (rdr.fieldnames or [])}
            key_col = cols.get("custref") or list(cols.values())[0]
            val_col = cols.get("primaryclientcode") or list(cols.values())[1]
            for row in rdr:
                read += 1
                k = soft_clean(row.get(key_col, ""))
                v = soft_clean(row.get(val_col, ""))
                if k and v:
                    self.client_map[k] = v
                    mapped += 1
        return read, mapped

    def export_csv(self, path: Path):
        if not self.rows:
            messagebox.showerror("Export", "No data to export.")
            return
        with path.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            # write display labels
            w.writerow([display_label(c) for c in self.columns])
            for r in self.rows:
                w.writerow([r.get(c, "") for c in self.columns])
        messagebox.showinfo("Export", f"Saved CSV to:\n{path}")

    def export_xlsx(self, path: Path):
        if not self.rows:
            messagebox.showerror("Export", "No data to export.")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except Exception:
            messagebox.showwarning(
                "Export", "openpyxl not installed; exporting CSV instead.")
            return self.export_csv(path.with_suffix(".csv"))
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Rows"
        # header with display labels
        header_labels = [display_label(c) for c in self.columns]
        ws.append(list(header_labels))
        for r in self.rows:
            ws.append([r.get(c, "") for c in self.columns])
        for i, col in enumerate(self.columns, 1):
            maxlen = max([len(str(display_label(col)))] +
                         [len(str(r.get(col, ""))) for r in self.rows])
            ws.column_dimensions[get_column_letter(
                i)].width = min(max(12, maxlen + 2), 60)
        wb.save(path)
        messagebox.showinfo("Export", f"Saved Excel to:\n{path}")

    def run_analyze(self, path_entry: str):
        self.clear_table()
        self.rows = []
        if not path_entry:
            messagebox.showerror("Input", "Choose a file or folder.")
            return
        self.clear_table()
        self.rows = []
        if not path_entry:
            messagebox.showerror("Input", "Choose a file or folder.")
            return

        p = Path(path_entry)
        files: List[Path] = []
        if p.is_file():
            files = [p]
        elif p.is_dir():
            files = [f for f in p.iterdir() if f.is_file()
                     and f.suffix.lower() == ".pdf"]
        else:
            messagebox.showerror("Input", "Path not found.")
            return

        # Local-only: no Azure client
        total_rows = 0
        fedex_count = 0
        lightning_count = 0
        # (for files that weren’t detected as FedEx or Lightning)
        generic_count = 0
        errors: List[str] = []

        total_files = len(files)
        self.set_progress(0, max(1, total_files))
        done = 0

        for f in files:
            try:
                # Always local
                txt = read_pdf_text(f)   # OCR fallback
                if looks_like_fedex(txt):
                    fedex_count += 1
                elif looks_like_lightning(txt):
                    lightning_count += 1
                else:
                    # Will still try both local parsers
                    generic_count += 1

                rows = process_file_auto(f, client_map=self.client_map)
                self.rows.extend(rows)
                total_rows += len(rows)

            except Exception as ex:
                errors.append(f"{f.name}: {ex}")

            # progress update per file
            done += 1
            self.set_progress(done, max(1, total_files))
            self.after_call(1, lambda: None)

        # Show rows
        for r in self.rows:
            self.add_row([r.get(c, "") for c in self.columns])

        # Status summary
        inv_totals: Dict[str, float] = {}
        for r in self.rows:
            inv = r.get("InvoiceID") or ""
            amt = r.get("Amount")
            if isinstance(amt, (int, float)) or (isinstance(amt, str) and amt):
                try:
                    a = float(amt) if not isinstance(
                        amt, (int, float)) else float(amt)
                except Exception:
                    a = 0.0
                inv_totals[inv] = inv_totals.get(inv, 0.0) + a

        msg = (f"Done. Files: {len(files)} FedEx: {fedex_count} "
               f"Lightning: {lightning_count} Other(local): {generic_count} Rows: {total_rows}")

        if inv_totals:
            joined = "; ".join(f"{k}=${v:,.2f}" for k,
                               v in inv_totals.items() if k)
            if joined:
                msg += " Totals (per InvoiceID): " + joined

        if errors:
            msg += f" Errors: {len(errors)} (see details)"
            # quick dialog with first few errors
            messagebox.showwarning("Some files failed",
                                   "\n".join(errors[:3]) + ("\n..." if len(errors) > 3 else ""))


# --------- CTk GUI ----------
if USE_CTK:
    class AppCTK(ctk.CTk, AppBase):
        def safe_destroy(self):
            """Safely destroy the window and cancel any repeating callbacks."""
            try:
                self._destroyed = True
            except Exception:
                pass
            try:
                super().destroy()
            except Exception:
                pass

        def after_call(self, ms, func):
            """Schedule a callback, but skip if window is destroyed."""
            def wrapper(*args, **kwargs):
                if getattr(self, '_destroyed', False):
                    return
                try:
                    func(*args, **kwargs)
                except Exception:
                    pass
            try:
                if not getattr(self, '_destroyed', False):
                    self.after(ms, wrapper)
            except Exception:
                pass

        def protocol_handlers(self):
            self.protocol("WM_DELETE_WINDOW", self.safe_destroy)

        def __init__(self):
            ctk.CTk.__init__(self)
            AppBase.__init__(self)
            ctk.set_appearance_mode("System")
            ctk.set_default_color_theme("blue")
            super().__init__()
            self.title(f"Smart Invoice Runner [{APP_VERSION}]")
            self.geometry("1280x900")
            self.minsize(1180, 820)
            # Set window icon
            try:
                self.iconbitmap(
                    r"C:\Users\rscottdeperto\Desktop\Invoice Testing\Coding\assets\icon64 (1).ico")
            except Exception:
                pass
            # Splash first
            self.show_launch_splash()
            self._build_left()
            self._build_right()
            # (Removed duplicate icon from main window)
            # close splash shortly after UI is ready
            self.after(1600, self.hide_launch_splash)

            # Add protocol handler for safe destroy
            self._destroyed = False
            self.protocol_handlers()

        def _build_left(self):
            left = ctk.CTkFrame(self, corner_radius=0, width=310)
            left.grid(row=0, column=0, sticky="ns")
            left.grid_rowconfigure(7, weight=1)
            # Add icon at top left, with transparent background
            try:
                icon_img = None
                if PIL_OK:
                    from PIL import Image, ImageTk
                    icon_img = Image.open(
                        r"C:\Users\rscottdeperto\Desktop\Invoice Testing\Coding\assets\icon64 (1).ico")
                    icon_img = icon_img.resize((40, 40), Image.LANCZOS)
                    # Convert to RGBA and set alpha for transparency
                    icon_img = icon_img.convert("RGBA")
                    datas = icon_img.getdata()
                    newData = []
                    for item in datas:
                        # If pixel is white or near-white, make it transparent
                        if item[0] > 240 and item[1] > 240 and item[2] > 240:
                            newData.append((255, 255, 255, 0))
                        else:
                            newData.append(item)
                    icon_img.putdata(newData)
                    icon_img = ImageTk.PhotoImage(icon_img)
                if icon_img:
                    self.left_icon_label = ctk.CTkLabel(
                        left, image=icon_img, text="")
                    self.left_icon_label.image = icon_img
                    self.left_icon_label.grid(
                        row=0, column=0, padx=12, pady=(12, 2), sticky="nw")
            except Exception:
                pass
            # Add bold/underlined header
            header_font = ctk.CTkFont(weight="bold", underline=True, size=15)
            ctk.CTkLabel(left, text="How to use this tool", font=header_font, justify="left").grid(
                row=1, column=0, padx=12, pady=(18, 2), sticky="w")
            # Instructions body
            instructions_body = instructions_text().split(
                '\n', 1)[1] if '\n' in instructions_text() else instructions_text()
            ctk.CTkLabel(left, text=instructions_body, justify="left", wraplength=270).grid(
                row=2, column=0, padx=12, pady=(2, 12), sticky="w")
            ctk.CTkLabel(left, text=f"App: {APP_VERSION}", justify="left").grid(
                row=7, column=0, padx=12, pady=(8, 10), sticky="sw")

        def _build_right(self):
            self.grid_rowconfigure(0, weight=1)
            self.grid_columnconfigure(1, weight=1)
            right = ctk.CTkFrame(self, corner_radius=0)
            right.grid(row=0, column=1, sticky="nsew")
            for c in range(12):
                right.grid_columnconfigure(c, weight=1)
            right.grid_rowconfigure(6, weight=1)

            # Top controls
            strip = ctk.CTkFrame(right)
            strip.grid(row=0, column=0, columnspan=12,
                       sticky="we", padx=10, pady=(10, 6))
            strip.grid_columnconfigure(3, weight=1)
            ctk.CTkLabel(strip, text="File/Folder:").grid(row=1,
                                                          column=0, padx=(6, 4), pady=6, sticky="e")
            self.var_path = tk.StringVar()
            ctk.CTkEntry(strip, textvariable=self.var_path, width=740).grid(
                row=1, column=1, columnspan=3, padx=(0, 8), pady=6, sticky="we")
            ctk.CTkButton(strip, text="Browse File", width=110, command=self._browse_file).grid(
                row=1, column=4, padx=(0, 6), pady=6, sticky="e")
            ctk.CTkButton(strip, text="Browse Folder", width=120, command=self._browse_folder).grid(
                row=1, column=5, padx=(0, 6), pady=6, sticky="e")
            ctk.CTkButton(strip, text="Analyze", width=130, command=self._analyze).grid(
                row=1, column=6, padx=(0, 8), pady=6, sticky="e")

            # Client Code Map (CSV)
            map_row = ctk.CTkFrame(right)
            map_row.grid(row=1, column=0, columnspan=12,
                         sticky="we", padx=10, pady=(0, 4))
            map_row.grid_columnconfigure(1, weight=1)
            ctk.CTkLabel(map_row, text="Client Code Map (CSV):").grid(
                row=0, column=0, padx=(6, 4), pady=6, sticky="e")
            self.var_csv = tk.StringVar()
            ctk.CTkEntry(map_row, textvariable=self.var_csv, width=740).grid(
                row=0, column=1, padx=(0, 8), pady=6, sticky="we")
            ctk.CTkButton(map_row, text="Browse CSV", width=110, command=self._browse_csv).grid(
                row=0, column=2, padx=(0, 6), pady=6, sticky="e")
            ctk.CTkButton(map_row, text="Load Map", width=110, command=self._load_map).grid(
                row=0, column=3, padx=(0, 6), pady=6, sticky="e")

            # Status + export
            status = ctk.CTkFrame(right)
            status.grid(row=2, column=0, columnspan=12,
                        sticky="we", padx=10, pady=(4, 8))
            status.grid_columnconfigure(0, weight=1)
            self.lbl_status = ctk.CTkLabel(status, text="Ready.")
            self.lbl_status.grid(row=0, column=0, padx=10, pady=6, sticky="w")

            # Graphical progress bar
            self.pbar = ctk.CTkProgressBar(status)
            self.pbar.set(0.0)
            self.pbar.grid(row=1, column=0, columnspan=2,
                           padx=10, pady=(0, 8), sticky="we")

            ctk.CTkButton(status, text="Export Excel", width=130, command=self._export_xlsx).grid(
                row=0, column=1, padx=6, pady=6, sticky="e")
            ctk.CTkButton(status, text="Export CSV", width=120, command=self._export_csv).grid(
                row=0, column=2, padx=(0, 10), pady=6, sticky="e")

            # Table
            self.tbl_frame = ctk.CTkFrame(right)
            self.tbl_frame.grid(row=6, column=0, columnspan=12,
                                sticky="nsew", padx=10, pady=(0, 10))
            self.tbl_frame.grid_rowconfigure(0, weight=1)
            self.tbl_frame.grid_columnconfigure(0, weight=1)
            style = ttk.Style(self.tbl_frame)
            style.theme_use('default')
            self.tree = None
            self.rebuild_tree()

        # Bridges
        def set_status(self, msg: str):
            self.lbl_status.configure(text=msg)

        def set_progress(self, done: int, total: int):
            try:
                frac = 0.0 if total <= 0 else max(0.0, min(1.0, done / total))
                self.pbar.set(frac)
            except Exception:
                pass

        def rebuild_tree(self):
            if getattr(self, "tree", None) is not None:
                try:
                    self.tree.destroy()
                except Exception:
                    pass
            self.tree = ttk.Treeview(
                self.tbl_frame, columns=COLUMNS_UNIFIED, show="headings")
            for c in COLUMNS_UNIFIED:
                width = 200 if c in ("Description", "InvoiceFileName") else 140
                if c in ("FedEx_Sender", "FedEx_CustRef", "PrimaryClientCode"):
                    width = 180
                self.tree.heading(c, text=display_label(c))
                self.tree.column(c, width=width, anchor="w")
            yscroll = ttk.Scrollbar(
                self.tbl_frame, orient="vertical", command=self.tree.yview)
            self.tree.configure(yscrollcommand=yscroll.set)
            self.tree.grid(row=0, column=0, sticky="nsew")
            yscroll.grid(row=0, column=1, sticky="ns")

        def clear_table(self):
            for iid in self.tree.get_children():
                self.tree.delete(iid)

        def add_row(self, values: List):
            self.tree.insert("", "end", values=values)

        def get_tree_column_width(self, name):
            return self.tree.column(name, option="width")

        def set_tree_rowheight(self, px):
            ttk.Style(self.tree).configure("Treeview", rowheight=int(px))

        def after_call(self, ms, func):
            self.after(ms, func)

        # ----- overlays -----
        def show_status_bubble(self, msg: str):
            if self._status_bubble is not None:
                try:
                    self._status_bubble.destroy()
                except Exception:
                    pass
            self._status_bubble = tk.Toplevel(self)
            try:
                self._status_bubble.iconbitmap(
                    r"C:\Users\rscottdeperto\Desktop\Invoice Testing\Coding\assets\icon64 (1).ico")
            except Exception:
                pass
            self._status_bubble.overrideredirect(True)
            self._status_bubble.attributes("-topmost", True)
            frm = ctk.CTkFrame(self._status_bubble)
            frm.pack(fill="both", expand=True, padx=8, pady=8)
            ctk.CTkLabel(frm, text=msg, font=ctk.CTkFont(
                size=14, weight="bold")).pack(padx=12, pady=(10, 6))
            pb = ttk.Progressbar(frm, mode="indeterminate",
                                 length=220, maximum=100)
            pb.pack(padx=12, pady=(0, 12))
            pb.start(12)
            # center near parent
            self.update_idletasks()
            x = self.winfo_x() + (self.winfo_width() // 2) - 140
            y = self.winfo_y() + 120
            self._status_bubble.geometry(f"+{x}+{y}")

        def hide_status_bubble(self):
            if self._status_bubble is not None:
                try:
                    self._status_bubble.destroy()
                except Exception:
                    pass
                self._status_bubble = None

        def _load_image_from_url(self, url: str, max_size=(560, 300)):
            if not PIL_OK:
                return None
            try:
                # If the path is a local file, open directly
                if url.lower().endswith('.png') or url.lower().endswith('.jpg') or url.lower().endswith('.jpeg'):
                    img = Image.open(url).convert("RGBA")
                    img.thumbnail(max_size, Image.LANCZOS)
                    return img
                # Otherwise, treat as URL
                r = requests.get(url, timeout=8)
                r.raise_for_status()
                img = Image.open(BytesIO(r.content)).convert("RGBA")
                img.thumbnail(max_size, Image.LANCZOS)
                return img
            except Exception:
                return None

        def show_launch_splash(self):
            if self._splash is not None:
                return
            self._splash = tk.Toplevel(self)
            try:
                self._splash.iconbitmap(
                    r"C:\Users\rscottdeperto\Desktop\Invoice Testing\Coding\assets\icon64 (1).ico")
            except Exception:
                pass
            self._splash.overrideredirect(True)
            self._splash.attributes("-topmost", True)
            frm = ctk.CTkFrame(self._splash, corner_radius=12)
            frm.pack(fill="both", expand=True, padx=10, pady=10)
            img = self._load_image_from_url(SPLASH_IMAGE_URL)
            if img is not None:
                cimg = ctk.CTkImage(
                    light_image=img, dark_image=img, size=img.size)
                ctk.CTkLabel(frm, image=cimg, text="").pack(padx=14, pady=14)
                # hold reference
                self._splash._img_ref = cimg
            else:
                ctk.CTkLabel(frm, text="Loading…", font=ctk.CTkFont(
                    size=18, weight="bold")).pack(padx=20, pady=20)
            self.update_idletasks()
            # center on screen
            sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
            w, h = 640, 360
            x, y = (sw - w)//2, (sh - h)//2
            self._splash.geometry(f"{w}x{h}+{x}+{y}")

        def hide_launch_splash(self):
            if self._splash is not None:
                try:
                    self._splash.destroy()
                except Exception:
                    pass
                self._splash = None

        # Actions
        def _browse_file(self):
            p = filedialog.askopenfilename(
                filetypes=[("PDF", "*.pdf"), ("All", "*.*")]
            )
            if p:
                self.var_path.set(p)

        def _browse_folder(self):
            p = filedialog.askdirectory()
            if p:
                self.var_path.set(p)

        def _browse_csv(self):
            p = filedialog.askopenfilename(
                filetypes=[("CSV", "*.csv"), ("All", "*.*")]
            )
            if p:
                self.var_csv.set(p)

        def _load_map(self):
            p = Path(self.var_csv.get().strip()
                     ) if self.var_csv.get().strip() else None
            if not p or not p.exists():
                messagebox.showerror("CSV", "Select a valid CSV first.")
                return
            try:
                read, mapped = self.load_client_map_csv(p)
                self.set_status(f"Loaded map: rows={read}, mapped={mapped}")
                messagebox.showinfo("Load Map", "Load Complete")
            except Exception as ex:
                messagebox.showerror("CSV", f"Failed to load: {ex}")

        def _export_xlsx(self):
            path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[
                                                    ("Excel Workbook", "*.xlsx")],
                                                initialfile="invoice_rows.xlsx")
            if not path:
                return
            self.export_xlsx(Path(path))

        def _export_csv(self):
            path = filedialog.asksaveasfilename(defaultextension=".csv",
                                                filetypes=[("CSV", "*.csv")],
                                                initialfile="invoice_rows.csv")
            if not path:
                return
            self.export_csv(Path(path))

        def _analyze(self):
            self.set_status("Analyzing…")
            self.show_status_bubble("Analyzing… Please wait")
            self.update_idletasks()
            try:
                self.run_analyze(self.var_path.get().strip())
            finally:
                self.hide_status_bubble()


# --------- Tk fallback ----------


class AppTk(tk.Tk, AppBase):
    def __init__(self):
        tk.Tk.__init__(self)
        AppBase.__init__(self)
        super().__init__()
        self.title(f"Smart Invoice Runner [{APP_VERSION}]")
        self.geometry("1220x880")
        self.minsize(1120, 800)
        # Set window icon
        try:
            self.iconbitmap(
                r"C:\Users\rscottdeperto\Desktop\Invoice Testing\Coding\assets\icon64 (1).ico")
        except Exception:
            pass
        # Splash first
        self.show_launch_splash()
        self._build_left()
        self._build_right()
        # Add icon to top left of main app (Tk)
        try:
            icon_img = None
            if PIL_OK:
                from PIL import Image, ImageTk
                icon_img = Image.open(
                    r"C:\Users\rscottdeperto\Desktop\Invoice Testing\Coding\assets\icon64 (1).ico")
                icon_img = icon_img.resize((40, 40), Image.LANCZOS)
                icon_img = ImageTk.PhotoImage(icon_img)
            if icon_img:
                self.icon_label = tk.Label(self, image=icon_img, bg="#f0f0f0")
                self.icon_label.image = icon_img
                self.icon_label.place(x=8, y=8)
        except Exception:
            pass
        self.after(1600, self.hide_launch_splash)

    def _build_left(self):
        left = tk.Frame(self, bg="#f0f0f0", width=320)
        left.grid(row=0, column=0, sticky="ns")
        left.grid_propagate(False)
        # Add icon at top left, with transparent background
        try:
            icon_img = None
            if PIL_OK:
                from PIL import Image, ImageTk
                icon_img = Image.open(
                    r"C:\Users\rscottdeperto\Desktop\Invoice Testing\Coding\assets\icon64 (1).ico")
                icon_img = icon_img.resize((40, 40), Image.LANCZOS)
                icon_img = icon_img.convert("RGBA")
                datas = icon_img.getdata()
                newData = []
                for item in datas:
                    if item[0] > 240 and item[1] > 240 and item[2] > 240:
                        newData.append((255, 255, 255, 0))
                    else:
                        newData.append(item)
                icon_img.putdata(newData)
                icon_img = ImageTk.PhotoImage(icon_img)
            if icon_img:
                self.left_icon_label = tk.Label(
                    left, image=icon_img, bg="#f0f0f0")
                self.left_icon_label.image = icon_img
                self.left_icon_label.pack(anchor="nw", padx=12, pady=(12, 2))
        except Exception:
            pass
        tk.Label(left, text="Smart Invoice Runner", font=(
            "Segoe UI", 12, "bold"), bg="#f0f0f0").pack(anchor="w", padx=12, pady=(4, 6))
        # Add bold/underlined header
        header_font = ("Segoe UI", 12, "bold")
        tk.Label(left, text="How to use this tool", font=header_font,
                 bg="#f0f0f0", underline=1).pack(anchor="w", padx=12, pady=(18, 2))
        # Instructions body
        instructions_body = instructions_text().split(
            '\n', 1)[1] if '\n' in instructions_text() else instructions_text()
        tk.Label(left, text=instructions_body, justify="left", wraplength=280,
                 bg="#f0f0f0").pack(anchor="w", padx=12, pady=(2, 10))
        tk.Label(left, text=f"App: {APP_VERSION}", bg="#f0f0f0").pack(
            side="bottom", anchor="w", padx=12, pady=(8, 10))

    def _build_right(self):
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        right = tk.Frame(self)
        right.grid(row=0, column=1, sticky="nsew")
        for c in range(12):
            right.grid_columnconfigure(c, weight=1)
        right.grid_rowconfigure(6, weight=1)

        strip = tk.Frame(right)
        strip.grid(row=0, column=0, columnspan=12,
                   sticky="we", padx=10, pady=(10, 6))
        tk.Label(strip, text="API Key (for non‑FedEx/Lightning):").grid(
            row=0, column=2, padx=(6, 4), pady=6, sticky="e")
        self.var_apikey = tk.StringVar()
        tk.Entry(strip, textvariable=self.var_apikey, show="•", width=40).grid(
            row=0, column=3, padx=(0, 8), pady=6, sticky="w")
        tk.Label(strip, text="API Ver:").grid(
            row=0, column=4, padx=(6, 4), pady=6, sticky="e")
        tk.Label(strip, text="File/Folder:").grid(row=1,
                                                  column=0, padx=(6, 4), pady=6, sticky="e")
        self.var_path = tk.StringVar()
        tk.Entry(strip, textvariable=self.var_path, width=80).grid(
            row=1, column=1, columnspan=3, padx=(0, 8), pady=6, sticky="we")
        tk.Button(strip, text="Browse File", width=12, command=self._browse_file).grid(
            row=1, column=4, padx=(0, 6), pady=6, sticky="e")
        tk.Button(strip, text="Browse Folder", width=12, command=self._browse_folder).grid(
            row=1, column=5, padx=(0, 6), pady=6, sticky="e")
        tk.Button(strip, text="Analyze", width=12, command=self._analyze).grid(
            row=1, column=6, padx=(0, 8), pady=6, sticky="e")

        # Client Code Map (CSV)
        map_row = tk.Frame(right)
        map_row.grid(row=1, column=0, columnspan=12,
                     sticky="we", padx=10, pady=(0, 4))
        map_row.grid_columnconfigure(1, weight=1)
        tk.Label(map_row, text="Client Code Map (CSV):").grid(
            row=0, column=0, padx=(6, 4), pady=6, sticky="e")
        self.var_csv = tk.StringVar()
        tk.Entry(map_row, textvariable=self.var_csv, width=80).grid(
            row=0, column=1, padx=(0, 8), pady=6, sticky="we")
        tk.Button(map_row, text="Browse CSV", width=12, command=self._browse_csv).grid(
            row=0, column=2, padx=(0, 6), pady=6, sticky="e")
        tk.Button(map_row, text="Load Map", width=12, command=self._load_map).grid(
            row=0, column=3, padx=(0, 6), pady=6, sticky="e")

        status = tk.Frame(right)
        status.grid(row=2, column=0, columnspan=12,
                    sticky="we", padx=10, pady=(4, 8))
        status.grid_columnconfigure(0, weight=1)
        self.lbl_status = tk.Label(status, text="Ready.")
        self.lbl_status.grid(row=0, column=0, padx=10, pady=6, sticky="w")

        # Graphical progress bar (ttk)
        self.pbar = ttk.Progressbar(
            status, mode="determinate", length=360, maximum=100)
        self.pbar["value"] = 0
        self.pbar.grid(row=1, column=0, columnspan=2,
                       padx=10, pady=(0, 8), sticky="w")

        tk.Button(status, text="Export Excel", width=14, command=self._export_xlsx).grid(
            row=0, column=1, padx=6, pady=6, sticky="e")
        tk.Button(status, text="Export CSV", width=12, command=self._export_csv).grid(
            row=0, column=2, padx=(0, 10), pady=6, sticky="e")

        self.tbl_frame = tk.Frame(right)
        self.tbl_frame.grid(row=6, column=0, columnspan=12,
                            sticky="nsew", padx=10, pady=(0, 10))
        self.tbl_frame.grid_rowconfigure(0, weight=1)
        self.tbl_frame.grid_columnconfigure(0, weight=1)
        self.tree = None
        self.rebuild_tree()

    # ---- Bridges (inside class) ----
    def set_status(self, msg: str):
        self.lbl_status.config(text=msg)

    def set_progress(self, done: int, total: int):
        try:
            pct = 0 if total <= 0 else int(
                max(0, min(100, round(100 * done / total))))
            self.pbar["value"] = pct
            self.update_idletasks()
        except Exception:
            pass

    def rebuild_tree(self):
        if getattr(self, "tree", None) is not None:
            try:
                self.tree.destroy()
            except Exception:
                pass
        self.tree = ttk.Treeview(
            self.tbl_frame, columns=COLUMNS_UNIFIED, show="headings")
        for c in COLUMNS_UNIFIED:
            width = 200 if c in ("Description", "InvoiceFileName") else 140
            if c in ("FedEx_Sender", "FedEx_CustRef", "PrimaryClientCode"):
                width = 180
            self.tree.heading(c, text=display_label(c))
            self.tree.column(c, width=width, anchor="w")
        yscroll = ttk.Scrollbar(
            self.tbl_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=yscroll.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")

    def clear_table(self):
        for iid in self.tree.get_children():
            self.tree.delete(iid)

    def add_row(self, values: List):
        self.tree.insert("", "end", values=values)

    def get_tree_column_width(self, name):
        return self.tree.column(name, option="width")

    def set_tree_rowheight(self, px):
        ttk.Style(self.tree).configure("Treeview", rowheight=int(px))

    def after_call(self, ms, func):
        self.after(ms, func)

    # ----- overlays -----
    def show_status_bubble(self, msg: str):
        if self._status_bubble is not None:
            try:
                self._status_bubble.destroy()
            except Exception:
                pass
        self._status_bubble = tk.Toplevel(self)
        try:
            self._status_bubble.iconbitmap(
                r"C:\Users\rscottdeperto\Desktop\Invoice Testing\Coding\assets\icon64 (1).ico")
        except Exception:
            pass
        self._status_bubble.overrideredirect(True)
        self._status_bubble.attributes("-topmost", True)
        frm = tk.Frame(self._status_bubble, bd=1, relief="ridge")
        frm.pack(fill="both", expand=True)
        tk.Label(frm, text=msg, font=(
            "Segoe UI", 11, "bold"), padx=14, pady=8).pack()
        pb = ttk.Progressbar(frm, mode="indeterminate",
                             length=220, maximum=100)
        pb.pack(padx=10, pady=(0, 10))
        pb.start(12)
        self.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - 140
        y = self.winfo_y() + 120
        self._status_bubble.geometry(f"+{x}+{y}")

    def hide_status_bubble(self):
        if self._status_bubble is not None:
            try:
                self._status_bubble.destroy()
            except Exception:
                pass
            self._status_bubble = None

    def _load_image_from_url(self, url: str, max_size=(560, 300)):
        if not PIL_OK:
            return None
        try:
            # If the path is a local file, open directly
            if url.lower().endswith('.png') or url.lower().endswith('.jpg') or url.lower().endswith('.jpeg'):
                img = Image.open(url).convert("RGBA")
                img.thumbnail(max_size, Image.LANCZOS)
                return img
            # Otherwise, treat as URL
            r = requests.get(url, timeout=8)
            r.raise_for_status()
            img = Image.open(BytesIO(r.content)).convert("RGBA")
            img.thumbnail(max_size, Image.LANCZOS)
            return img
        except Exception:
            return None

    def show_launch_splash(self):
        if self._splash is not None:
            return
        self._splash = tk.Toplevel(self)
        try:
            self._splash.iconbitmap(
                r"C:\Users\rscottdeperto\Desktop\Invoice Testing\Coding\assets\icon64 (1).ico")
        except Exception:
            pass
        self._splash.overrideredirect(True)
        self._splash.attributes("-topmost", True)
        frm = tk.Frame(self._splash, bd=1, relief="ridge")
        frm.pack(fill="both", expand=True, padx=6, pady=6)
        img = self._load_image_from_url(SPLASH_IMAGE_URL)
        if img is not None:
            imgtk = ImageTk.PhotoImage(img)
            lbl = tk.Label(frm, image=imgtk)
            lbl.image = imgtk
            lbl.pack(padx=10, pady=10)
        else:
            tk.Label(frm, text="Loading…", font=(
                "Segoe UI", 14, "bold")).pack(padx=16, pady=16)
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w, h = 640, 360
        x, y = (sw - w)//2, (sh - h)//2
        self._splash.geometry(f"{w}x{h}+{x}+{y}")

    def hide_launch_splash(self):
        if self._splash is not None:
            try:
                self._splash.destroy()
            except Exception:
                pass
            self._splash = None

    # ---- Actions (inside class) ----
    def _browse_file(self):
        p = filedialog.askopenfilename(
            filetypes=[("PDF", "*.pdf"), ("All", "*.*")]
        )
        if p:
            self.var_path.set(p)

    def _browse_folder(self):
        p = filedialog.askdirectory()
        if p:
            self.var_path.set(p)

    def _browse_csv(self):
        p = filedialog.askopenfilename(
            filetypes=[("CSV", "*.csv"), ("All", "*.*")]
        )
        if p:
            self.var_csv.set(p)

    def _load_map(self):
        p = Path(self.var_csv.get().strip()
                 ) if self.var_csv.get().strip() else None
        if not p or not p.exists():
            messagebox.showerror("CSV", "Select a valid CSV first.")
            return
        try:
            read, mapped = self.load_client_map_csv(p)
            self.set_status(f"Loaded map: rows={read}, mapped={mapped}")
            messagebox.showinfo("Load Map", "Load Complete")
        except Exception as ex:
            messagebox.showerror("CSV", f"Failed to load: {ex}")

    def _export_xlsx(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[
                                                ("Excel Workbook", "*.xlsx")],
                                            initialfile="invoice_rows.xlsx")
        if not path:
            return
        self.export_xlsx(Path(path))

    def _export_csv(self):
        path = filedialog.asksaveasfilename(defaultextension=".csv",
                                            filetypes=[("CSV", "*.csv")],
                                            initialfile="invoice_rows.csv")
        if not path:
            return
        self.export_csv(Path(path))

    def _analyze(self):
        self.set_status("Analyzing…")
        self.show_status_bubble("Analyzing… Please wait")
        self.update_idletasks()
        try:
            self.run_analyze(self.var_path.get().strip())
        finally:
            self.hide_status_bubble()


# ======================================
# Entry
# ======================================


def main():
    app = AppCTK() if USE_CTK else AppTk()
    try:
        app.mainloop()
    except (KeyboardInterrupt, Exception) as e:
        import tkinter
        if isinstance(e, tkinter.TclError):
            pass  # Suppress TclError on exit
        elif isinstance(e, KeyboardInterrupt):
            pass  # Suppress Ctrl+C
        else:
            raise


if __name__ == "__main__":
    main()
